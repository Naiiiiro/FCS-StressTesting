import pandas as pd
from openpyxl.styles import Alignment

def calculate_expected_loss(ead, pd_rate, lgd_rate):
    el = pd_rate * lgd_rate * ead
    return pd_rate, lgd_rate, el

def process_sheet(sheet_name, df):
    total_ead = df['違約曝險額'].sum()
    return {
        '授信部位': sheet_name,
        '暴險金額': total_ead
    }

def save_to_excel(summary_df, sheet_name, writer, pd_rate, lgd_rate, pd_rate_orderly, lgd_rate_orderly, pd_rate_disorderly, lgd_rate_disorderly, pd_rate_no_policy, lgd_rate_no_policy):
    for index, row in summary_df.iterrows():
        ead = row['暴險金額']
        
        # 基準情境
        base_pd, base_lgd, base_el = calculate_expected_loss(ead, pd_rate, lgd_rate)
        summary_df.at[index, '基準情境_平均違約率'] = base_pd
        summary_df.at[index, '基準情境_平均違約損失率'] = base_lgd
        summary_df.at[index, '基準情境_估計可能損失數'] = base_el

        # 有序轉型
        orderly_pd, orderly_lgd, orderly_el = calculate_expected_loss(ead, pd_rate_orderly, lgd_rate_orderly)
        summary_df.at[index, '平均違約率'] = orderly_pd
        summary_df.at[index, '平均違約損失率'] = orderly_lgd
        summary_df.at[index, '估計可能損失數'] = orderly_el

        # 無序轉型
        disorderly_pd, disorderly_lgd, disorderly_el = calculate_expected_loss(ead, pd_rate_disorderly, lgd_rate_disorderly)
        summary_df.at[index, '平均違約率'] = disorderly_pd
        summary_df.at[index, '平均違約損失率'] = disorderly_lgd
        summary_df.at[index, '估計可能損失數'] = disorderly_el

        # 無政策情境
        no_policy_pd, no_policy_lgd, no_policy_el = calculate_expected_loss(ead, pd_rate_no_policy, lgd_rate_no_policy)
        summary_df.at[index, '平均違約率'] = no_policy_pd
        summary_df.at[index, '平均違約損失率'] = no_policy_lgd
        summary_df.at[index, '估計可能損失數'] = no_policy_el

    summary_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
    workbook = writer.book
    worksheet = workbook[sheet_name]

    worksheet.merge_cells('C1:E1')
    worksheet['C1'] = '基準情境'
    worksheet['C1'].alignment = Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells('F1:H1')
    worksheet['F1'] = '有序轉型'
    worksheet['F1'].alignment = Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells('I1:K1')
    worksheet['I1'] = '無序轉型'
    worksheet['I1'].alignment = Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells('L1:N1')
    worksheet['L1'] = '無政策情境'
    worksheet['L1'].alignment = Alignment(horizontal='center', vertical='center')

def main():
    file_path = '國內企業授信/國內企業授信-營建業.xlsx'
    xls = pd.ExcelFile(file_path)
    
    pd_rate_2030 = 0.02
    lgd_rate_2030 = 0.4
    pd_rate_2050 = 0.03
    lgd_rate_2050 = 0.5
    pd_rate_2100 = 0.03
    lgd_rate_2100 = 0.5

    pd_rate_orderly_2030 = 0.025
    lgd_rate_orderly_2030 = 0.45
    pd_rate_disorderly_2030 = 0.03
    lgd_rate_disorderly_2030 = 0.5
    pd_rate_no_policy_2030 = 0.035
    lgd_rate_no_policy_2030 = 0.55

    pd_rate_orderly_2050 = 0.035
    lgd_rate_orderly_2050 = 0.55
    pd_rate_disorderly_2050 = 0.04
    lgd_rate_disorderly_2050 = 0.6
    pd_rate_no_policy_2050 = 0.045
    lgd_rate_no_policy_2050 = 0.65

    pd_rate_orderly_2100 = 0.035
    lgd_rate_orderly_2100 = 0.55
    pd_rate_disorderly_2100 = 0.04
    lgd_rate_disorderly_2100 = 0.6
    pd_rate_no_policy_2100 = 0.045
    lgd_rate_no_policy_2100 = 0.65
    
    summary_2030 = []
    summary_2050 = []
    summary_2100 = []
    
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name)
        summary_2030.append(process_sheet(sheet_name, df))
        summary_2050.append(process_sheet(sheet_name, df))
        summary_2100.append(process_sheet(sheet_name, df))
    
    summary_df_2030 = pd.DataFrame(summary_2030)
    summary_df_2050 = pd.DataFrame(summary_2050)
    summary_df_2100 = pd.DataFrame(summary_2100)

    with pd.ExcelWriter('氣候風險壓力測試_結果.xlsx', engine='openpyxl') as writer:
        save_to_excel(summary_df_2030, '國內授信彙總表(2030年)', writer, pd_rate_2030, lgd_rate_2030, pd_rate_orderly_2030, lgd_rate_orderly_2030, pd_rate_disorderly_2030, lgd_rate_disorderly_2030, pd_rate_no_policy_2030, lgd_rate_no_policy_2030)
        save_to_excel(summary_df_2050, '國內授信彙總表(2050年)', writer, pd_rate_2050, lgd_rate_2050, pd_rate_orderly_2050, lgd_rate_orderly_2050, pd_rate_disorderly_2050, lgd_rate_disorderly_2050, pd_rate_no_policy_2050, lgd_rate_no_policy_2050)
        save_to_excel(summary_df_2100, '國內授信彙總表(2100年)', writer, pd_rate_2100, lgd_rate_2100, pd_rate_orderly_2100, lgd_rate_orderly_2100, pd_rate_disorderly_2100, lgd_rate_disorderly_2100, pd_rate_no_policy_2100, lgd_rate_no_policy_2100)
    
    print("結果已儲存至 氣候風險壓力測試_結果.xlsx")

if __name__ == "__main__":
    main()
